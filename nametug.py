import openpyxl
from tkinter import *

#LOAD AND ACTIVATE WORKBOOK (NOTE: REPLACE DIRETORY TO AN EXCEL WHERE YOUR EMAILS ARE STORED! BELOW IS JUST AN EXAMPLE...)
book = openpyxl.load_workbook(r'C:\Users\scheerio\Desktop\Book1.xlsx')
ws = book.active

#LOAD COMPLETE BOOLEAN
done = False

#ARRAYS TO HOLD FIRST AND LAST NAMES
first=[]
last=[]

#MAKE HEADERS
ws['B1']='First Name'
ws['C1']='Last Name'

#GET THE NAMES
firstone=True
middlepart=False
endpart=False
dotcount=0
counter=0
firstname=''
lastname=''
pathnumber=0
for cell in ws['A']:
	counter = counter + 1
	for x in cell.value:
		if (x=='.'):
			pathnumber=pathnumber+1
	#FOR 3-DOT EMAILS
	if (firstone==True):
			firstone=False
			continue
	if (pathnumber==3):
		if (cell.value!=None):
			for x in cell.value:
				if (x=='.'):
					dotcount=dotcount+1
				if (x=='@'):
					break
				if (dotcount==0):
					firstname=firstname+x
				if (dotcount==2 and x!='.'):
					lastname=lastname+x
			ws['B'+str(counter)]=firstname
			ws['C'+str(counter)]=lastname
			firstname=''
			lastname=''
			dotcount=0
	#FOR 2-DOT EMAILS
	if (pathnumber==2):
		if (cell.value!=None):
			for x in cell.value:
				if (x=='.'):
					dotcount=dotcount+1
				if (x=='@'):
					break
				if (dotcount==0):
					firstname=firstname+x
				if (dotcount==1 and x!='.'):
					lastname=lastname+x
			ws['B'+str(counter)]=firstname
			ws['C'+str(counter)]=lastname
			firstname=''
			lastname=''
			dotcount=0
	#FOR 1-DOT EMAILS
	if (pathnumber==1):
		if (cell.value!=None):
			for x in cell.value:
				if (x=='@'):
					break
				else:
					firstname=firstname+x
			ws['B'+str(counter)]=firstname
			firstname=''
			dotcount=0
	pathnumber=0

#WRAP IT UP
book.save('Nametug_Result.xlsx')
done=True
print('done')
		
#DISPLAY SUCCESS MESSAGE
if (done):
	root = Tk()
	labelfont = ('times', 20, 'bold')   
	root.title('Success Confirmation')               
	widget = Label(root, text='The excel file has been updated.', wraplength=600, justify=LEFT)
	widget.config(height=35, width=90)
	widget.pack(expand=YES, fill=BOTH)
	root.mainloop()












